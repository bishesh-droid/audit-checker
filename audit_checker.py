#!/usr/bin/env python3
"""
audit_checker.py
================
Coursera content-materials auditor.

Reads course rows from an Excel file, extracts the real Google Drive
hyperlinks hidden behind friendly cell labels in six asset columns
(Course Outline, PPTs, Written Assets, Final Videos, Raw Videos,
Course Artifacts Link), scans local hard drives for matching content,
optionally verifies Drive link accessibility, and produces a
colour-coded per-course / per-asset Excel report.

Python  : 3.10+
"""

from __future__ import annotations

import argparse
import json
import logging
import multiprocessing
import os
import pickle
import re
import shutil
import sys
import time
import urllib.error
import urllib.request
from concurrent.futures import ProcessPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from rapidfuzz import fuzz
from tqdm import tqdm

# ── Optional pydrive2 ─────────────────────────────────────────────────────────
try:
    from pydrive2.auth import GoogleAuth
    from pydrive2.drive import GoogleDrive
    GDRIVE_AVAILABLE = True
except ImportError:
    GDRIVE_AVAILABLE = False

# ── Optional gdown (folder downloads) ────────────────────────────────────────
try:
    import gdown
    GDOWN_AVAILABLE = True
except ImportError:
    GDOWN_AVAILABLE = False


# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

def setup_logging(log_file: str = "audit_checker.log", level: str = "INFO") -> logging.Logger:
    _logger = logging.getLogger("audit_checker")
    _logger.setLevel(getattr(logging, level.upper(), logging.INFO))
    _logger.handlers.clear()
    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)-8s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setFormatter(fmt)
    _logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stderr)
    ch.setFormatter(fmt)
    ch.setLevel(logging.WARNING)
    _logger.addHandler(ch)
    return _logger


logger = setup_logging()


# ═══════════════════════════════════════════════════════════════════════════════
#  ASSET COLUMN DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════════════

# Each entry: Excel column label → report key + local path-search hints.
# "hints" are lowercase keywords looked for anywhere in the local file path.
ASSET_COLUMNS: dict[str, dict] = {
    "Course Outline":               {"key": "Course_Outline",   "hints": ["outline", "syllabus", "cod"]},
    "PPTs":                         {"key": "PPTs",             "hints": ["ppt", "slide", "presentation"]},
    "Written Assets (PQ, GQ, DP)": {"key": "Written_Assets",   "hints": ["written", "pq", "gq", "quiz", "discussion"]},
    "Final Videos":                 {"key": "Final_Videos",     "hints": ["final", "video", "mp4", "mov", "mkv", "avi"]},
    "Raw Videos":                   {"key": "Raw_Videos",       "hints": ["raw", "footage", "rushes"]},
    "Course Artifacts Link":        {"key": "Course_Artifacts", "hints": []},
}


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA STRUCTURES
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class CourseEntry:
    """One course row parsed from the Excel file."""
    course_name: str
    semester:    str = ""
    term:        str = ""
    status:      str = ""
    source_file: str = ""
    row_index:   int = 0
    # Maps ASSET_COLUMNS column label → Google Drive URL (None if no hyperlink)
    asset_links: dict[str, Optional[str]] = field(default_factory=dict)


@dataclass
class AssetResult:
    """Audit outcome for one asset type within one course."""
    found_locally: str = "No"
    local_path:    str = ""
    drive_status:  str = "Not Checked"


@dataclass
class CourseAuditResult:
    """Full audit result for one course across all six asset types."""
    course_name:     str = ""
    semester:        str = ""
    term:            str = ""
    status:          str = ""
    # Maps ASSET_COLUMNS column label → AssetResult
    asset_results:   dict[str, AssetResult] = field(default_factory=dict)
    download_status: str = "Not Attempted"


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_CONFIG: dict = {
    "gsheet_url": "",          # Google Sheets URL — primary input source
    "excel_dir":  "./excel",   # Fallback: local .xlsx/.csv files directory
    "drives":     [],
    "output":     "./availability_report.xlsx",
    "columns": {
        "course_name": "Course",
        "semester":    "Sem",
        "term":        "Term",
        "status":      "Status",
    },
    "scanning": {
        "fuzzy_threshold":        75,
        "cache_file":             ".drive_index_cache.pkl",
        "cache_max_age_hours":    24,
        "gsheet_cache_hours":     1,   # Re-download sheet after this many hours
        "max_workers":            multiprocessing.cpu_count(),
        "extensions_filter":      [],
    },
    "google_drive": {
        "enabled":          False,
        "credentials_file": "credentials.json",
        "settings_file":    "settings.yaml",
        "min_free_gb":      5.0,
    },
    "logging": {
        "log_file": "audit_checker.log",
        "level":    "INFO",
    },
}


def _deep_merge(base: dict, override: dict) -> None:
    for key, value in override.items():
        if key in base and isinstance(base[key], dict) and isinstance(value, dict):
            _deep_merge(base[key], value)
        else:
            base[key] = value


def load_config(config_path: Optional[str] = None) -> dict:
    import copy
    config = copy.deepcopy(DEFAULT_CONFIG)
    candidates = [config_path] if config_path else ["config.json", "audit_config.json"]
    for candidate in candidates:
        if not candidate:
            continue
        p = Path(candidate)
        if p.exists():
            try:
                with open(p, encoding="utf-8") as fh:
                    _deep_merge(config, json.load(fh))
                logger.info("Loaded config from: %s", p)
                break
            except (json.JSONDecodeError, OSError) as exc:
                logger.warning("Could not load config '%s': %s", p, exc)
    return config


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL PARSING  —  reads all 6 asset columns + hyperlinks at once
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_course_file(
    filepath:   Path,
    course_col: str,
    sem_col:    str,
    term_col:   str,
    status_col: str,
) -> list[CourseEntry]:
    """
    Parse one .xlsx file with openpyxl so that hidden hyperlink URLs are
    captured from every ASSET_COLUMNS column alongside the cell display values.

    pandas.read_excel() only returns the display label (e.g. "Writing Practice")
    and silently discards the actual Drive URL stored as the hyperlink href.
    openpyxl exposes cell.hyperlink.target which contains the real URL.
    """
    entries: list[CourseEntry] = []
    try:
        wb = load_workbook(str(filepath), data_only=True, read_only=False)
        ws = wb.active

        # Build header-label → 1-based column index map
        col_idx: dict[str, int] = {
            str(cell.value or "").strip(): cell.column
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
            if cell.value
        }

        if course_col not in col_idx:
            logger.warning(
                "Column '%s' not found in '%s'. Available: %s",
                course_col, filepath.name, list(col_idx),
            )
            return []

        for row in ws.iter_rows(min_row=2):
            def _val(col: str) -> str:
                c = col_idx.get(col)
                return str(row[c - 1].value or "").strip() if c else ""

            name = _val(course_col)
            if not name or name.lower() in {"nan", "none", "n/a", ""}:
                continue

            # Collect the real hyperlink URL for every asset column
            asset_links: dict[str, Optional[str]] = {}
            for asset_col in ASSET_COLUMNS:
                c = col_idx.get(asset_col)
                if c:
                    cell = row[c - 1]
                    if cell.hyperlink and cell.hyperlink.target:
                        asset_links[asset_col] = cell.hyperlink.target.strip()

            entries.append(CourseEntry(
                course_name=name,
                semester=_val(sem_col),
                term=_val(term_col),
                status=_val(status_col),
                source_file=filepath.name,
                row_index=row[0].row,
                asset_links=asset_links,
            ))

    except Exception as exc:
        logger.error("Cannot parse '%s': %s", filepath, exc)

    return entries


def read_excel_courses(
    excel_dir:  str,
    course_col: str = "Course",
    sem_col:    str = "Sem",
    term_col:   str = "Term",
    status_col: str = "Status",
) -> list[CourseEntry]:
    """
    Discover all .xlsx / .csv files in *excel_dir* and parse course entries.
    Deduplicates by course name (keeps first occurrence).
    """
    dir_path = Path(excel_dir)
    if not dir_path.exists():
        raise FileNotFoundError(f"Excel directory not found: {excel_dir!r}")

    files = sorted(list(dir_path.rglob("*.xlsx")) + list(dir_path.rglob("*.csv")))
    if not files:
        logger.warning("No .xlsx/.csv files found in: %s", excel_dir)
        return []

    all_entries: list[CourseEntry] = []
    for fp in tqdm(files, desc="Reading Excel/CSV files", unit="file"):
        batch = _parse_course_file(fp, course_col, sem_col, term_col, status_col)
        links = sum(1 for e in batch if e.asset_links)
        logger.info("  %s → %d entries, %d with Drive links", fp.name, len(batch), links)
        all_entries.extend(batch)

    # Deduplicate — keep first occurrence of each course name
    seen:   set[str]          = set()
    unique: list[CourseEntry] = []
    for entry in all_entries:
        key = entry.course_name.lower()
        if key not in seen:
            seen.add(key)
            unique.append(entry)

    logger.info("Total unique courses: %d", len(unique))
    return unique


# ═══════════════════════════════════════════════════════════════════════════════
#  LOCAL DRIVE SCANNER
# ═══════════════════════════════════════════════════════════════════════════════

_SKIP_DIRS: frozenset[str] = frozenset({
    "System Volume Information", "$RECYCLE.BIN", "$Recycle.Bin",
    "Windows", "WinSxS", "SoftwareDistribution", "Recovery",
    "proc", "sys", "dev",
})


def _scan_single_drive(args: tuple[str, list[str]]) -> list[str]:
    """
    Recursively walk one drive and return all file + directory paths.
    Both are indexed so that course *folders* can be matched by name,
    not just individual files.
    """
    root_path, extensions_filter = args
    exts  = frozenset(e.lower().lstrip(".") for e in extensions_filter) if extensions_filter else frozenset()
    found: list[str] = []

    for dirpath, dirnames, filenames in os.walk(root_path, topdown=True, followlinks=False):
        dirnames[:] = [
            d for d in dirnames
            if not d.startswith(".") and d not in _SKIP_DIRS
        ]

        # Index directory names — course content is usually in a named folder
        for dname in dirnames:
            found.append(os.path.join(dirpath, dname))

        for fname in filenames:
            if exts and Path(fname).suffix.lower().lstrip(".") not in exts:
                continue
            found.append(os.path.join(dirpath, fname))

    return found


def build_file_index(
    drive_paths:         list[str],
    cache_file:          str   = ".drive_index_cache.pkl",
    cache_max_age_hours: float = 24.0,
    max_workers:         int   = 4,
    extensions_filter:   Optional[list[str]] = None,
) -> tuple[dict[str, list[str]], list[str]]:
    """
    Scan all drives and build:
      - ``filename_index``: dict[lowercase_name → [full_paths]]  (fast exact/fuzzy lookup)
      - ``all_paths``:      flat list of every path found         (used for path-based search)

    Both are persisted in a pickle cache.
    """
    cache_path = Path(cache_file)
    drives_key = sorted(drive_paths)

    if cache_path.exists():
        age_h = (time.time() - cache_path.stat().st_mtime) / 3600
        if age_h < cache_max_age_hours:
            try:
                with open(cache_path, "rb") as fh:
                    cached = pickle.load(fh)
                if cached.get("drives") == drives_key:
                    logger.info(
                        "Using cached index (%.1fh old, %d unique keys, %d total paths).",
                        age_h, len(cached["index"]), len(cached.get("all_paths", [])),
                    )
                    return cached["index"], cached.get("all_paths", [])
            except Exception as exc:
                logger.warning("Cache load failed (%s) — rescanning.", exc)

    all_paths: list[str] = []
    scan_args = [(d, extensions_filter or []) for d in drive_paths]
    workers   = max(1, min(max_workers, len(drive_paths)))

    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_scan_single_drive, arg): arg[0] for arg in scan_args}
        with tqdm(total=len(futures), desc="Scanning drives", unit="drive") as pbar:
            for future in as_completed(futures):
                drv = futures[future]
                try:
                    result = future.result()
                    all_paths.extend(result)
                    logger.info("  Drive '%s': %d paths.", drv, len(result))
                except Exception as exc:
                    logger.error("Scan failed for '%s': %s", drv, exc)
                finally:
                    pbar.update(1)

    # Build filename → [paths] index
    index: dict[str, list[str]] = {}
    for path in tqdm(all_paths, desc="Building index", unit="path", leave=False):
        p    = Path(path)
        full = p.name.lower()
        stem = p.stem.lower()
        for key in (full, stem):
            index.setdefault(key, []).append(path)

    try:
        with open(cache_path, "wb") as fh:
            pickle.dump({"drives": drives_key, "index": index, "all_paths": all_paths}, fh)
        logger.info("Index cached to: %s", cache_path)
    except Exception as exc:
        logger.warning("Could not write cache: %s", exc)

    return index, all_paths


# ═══════════════════════════════════════════════════════════════════════════════
#  PATH-BASED COURSE / ASSET SEARCH
# ═══════════════════════════════════════════════════════════════════════════════

def _score_path_for_course(course_lower: str, path: str) -> int:
    """
    Return the best similarity score between *course_lower* and any single
    path component (directory name or filename stem) in *path*.

    Uses both token_sort_ratio (good for word-reordered names) and
    partial_ratio (good when the folder name is a truncated prefix of the
    full course name, e.g. "Video Games" vs "Video Games - Technology…").
    """
    parts = re.split(r"[\\/]+", path.lower())
    best  = 0
    for part in parts:
        if len(part) < 3:
            continue
        s = max(
            fuzz.token_sort_ratio(course_lower, part),
            fuzz.partial_ratio(course_lower, part),
        )
        if s > best:
            best = s
    return best


def match_course_paths(
    course_name:     str,
    all_paths:       list[str],
    fuzzy_threshold: int = 75,
) -> list[tuple[str, int]]:
    """
    Find all local paths whose name fuzzy-matches *course_name*.

    Returns a list of ``(path, score)`` sorted by score descending.
    This is computed ONCE per course and reused for all six asset types.
    """
    course_lower = course_name.lower()
    candidates: list[tuple[str, int]] = []

    for path in all_paths:
        score = _score_path_for_course(course_lower, path)
        if score >= fuzzy_threshold:
            candidates.append((path, score))

    return sorted(candidates, key=lambda x: -x[1])


def find_asset_in_candidates(
    candidates:  list[tuple[str, int]],
    asset_hints: list[str],
) -> tuple[bool, str]:
    """
    Given pre-filtered course-matching paths, find the best one that also
    contains an asset-type keyword somewhere in its full path.

    Falls back to the best course-match path if no hint-specific path exists
    (the course folder is present even if the specific sub-folder isn't found).

    Args:
        candidates:  Output of :func:`match_course_paths` for this course.
        asset_hints: Lowercase keywords to look for in the full path string.

    Returns:
        ``(found, best_matching_path)``
    """
    if not candidates:
        return False, ""

    if not asset_hints:
        return True, candidates[0][0]

    # Prefer paths that also contain an asset keyword
    for path, _ in candidates:
        if any(hint in path.lower() for hint in asset_hints):
            return True, path

    # Course folder exists but specific asset sub-folder not identified
    return True, candidates[0][0]


# ═══════════════════════════════════════════════════════════════════════════════
#  GOOGLE DRIVE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

_DRIVE_ID_PATTERNS: list[str] = [
    r"/file/d/([a-zA-Z0-9_-]{10,})",
    r"/folders/([a-zA-Z0-9_-]{10,})",
    r"[?&]id=([a-zA-Z0-9_-]{10,})",
    r"/open\?id=([a-zA-Z0-9_-]{10,})",
    r"^([a-zA-Z0-9_-]{25,})$",
]


def extract_drive_file_id(link: str) -> Optional[str]:
    if not link:
        return None
    for pattern in _DRIVE_ID_PATTERNS:
        m = re.search(pattern, link.strip())
        if m:
            return m.group(1)
    return None


def authenticate_gdrive(
    credentials_file: str = "credentials.json",
    settings_file:    str = "settings.yaml",
) -> Optional[object]:
    if not GDRIVE_AVAILABLE:
        logger.warning("pydrive2 not installed — Google Drive API unavailable.")
        return None
    creds_path    = Path(credentials_file)
    settings_path = Path(settings_file)
    if not creds_path.exists() and not settings_path.exists():
        logger.warning("No Drive credentials found. Falling back to public HTTP checks.")
        return None
    try:
        gauth = GoogleAuth(
            settings_file=str(settings_path) if settings_path.exists() else None
        )
        saved = Path("mycreds.txt")
        if saved.exists():
            gauth.LoadCredentialsFile(str(saved))
        if gauth.credentials is None:
            gauth.LocalWebserverAuth()
        elif gauth.access_token_expired:
            gauth.Refresh()
        else:
            gauth.Authorize()
        gauth.SaveCredentialsFile(str(saved))
        logger.info("Google Drive authenticated.")
        return GoogleDrive(gauth)
    except Exception as exc:
        logger.error("GDrive auth failed: %s", exc)
        return None


def download_gsheet(
    url:         str,
    dest_dir:    str   = "./excel",
    cache_hours: float = 1.0,
    force:       bool  = False,
) -> Optional[str]:
    """
    Download a Google Sheets (or Drive-hosted .xlsx) file as .xlsx.

    The downloaded file is cached in *dest_dir*.  If the cached copy is
    younger than *cache_hours* hours it is reused without hitting the
    network (pass ``force=True`` or ``--no_cache`` to always re-download).

    The sheet must be publicly accessible ("Anyone with the link can view").
    Returns the local path to the saved file, or None on failure.
    """
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        logger.error("Cannot parse spreadsheet ID from URL: %s", url)
        print("      ERROR: Not a valid Google Sheets URL.")
        return None

    sheet_id   = m.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)
    out_path = dest / f"gsheet_{sheet_id}.xlsx"

    # ── Return cached copy if it is fresh enough ──────────────────────────────
    if not force and out_path.exists() and cache_hours > 0:
        age_h = (time.time() - out_path.stat().st_mtime) / 3600
        if age_h < cache_hours:
            print(f"      Using cached sheet ({age_h * 60:.0f} min old) → {out_path}")
            logger.info("Reusing cached Google Sheet: %s", out_path)
            return str(out_path)

    # ── Download fresh copy ───────────────────────────────────────────────────
    print(f"      Sheet ID : {sheet_id}")
    print(f"      Fetching : {export_url}")
    try:
        req = urllib.request.Request(export_url)
        req.add_header("User-Agent", "Mozilla/5.0 (compatible; AuditChecker/2.0)")
        with urllib.request.urlopen(req, timeout=60) as resp:
            ct = resp.headers.get("Content-Type", "")
            if "text/html" in ct:
                print("      ERROR: Sheet is private or login-protected.")
                print("      Share it as 'Anyone with the link can view' and retry.")
                return None
            data = resp.read()
        with open(out_path, "wb") as fh:
            fh.write(data)
        logger.info("Google Sheet saved: %s (%d bytes)", out_path, len(data))
        print(f"      Saved {len(data):,} bytes → {out_path}")
        return str(out_path)

    except urllib.error.HTTPError as exc:
        msgs = {403: "Access denied — share the sheet publicly.",
                404: "Sheet not found — check the URL."}
        print(f"      ERROR: HTTP {exc.code} — {msgs.get(exc.code, exc.reason)}")
        return None
    except Exception as exc:
        logger.error("GSheet download failed: %s", exc)
        print(f"      ERROR: {exc}")
        return None


def _public_gdrive_check(file_id: str, is_folder: bool = False) -> str:
    """
    Check whether a Google Drive file or folder is publicly accessible.

    Uses the folder-view URL for folders and the file-download URL for files.
    Detects login-page redirects (private content) by inspecting the final URL.
    """
    url = (
        f"https://drive.google.com/drive/folders/{file_id}"
        if is_folder
        else f"https://drive.google.com/uc?id={file_id}&export=download"
    )
    try:
        req = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0 (AuditChecker/1.0)")
        with urllib.request.urlopen(req, timeout=15) as resp:
            final_url = resp.geturl()
            # Google redirects to its login page for private/restricted content
            if "accounts.google.com" in final_url or "ServiceLogin" in final_url:
                return "Missing"
            if resp.status == 200:
                return "Available"
            return "Broken Link"
    except urllib.error.HTTPError as exc:
        return "Missing" if exc.code in (403, 404) else "Broken Link"
    except Exception:
        return "Broken Link"


def check_gdrive_link(
    link:  Optional[str],
    drive: Optional[object] = None,
) -> str:
    """
    Return a Drive status string for *link*:
    ``"Available"``, ``"Missing"``, ``"Broken Link"``, ``"No Link"``, or ``"Not Checked"``.
    """
    if not link:
        return "No Link"
    file_id = extract_drive_file_id(link)
    if not file_id:
        return "Broken Link"
    is_folder = "/folders/" in link
    if drive is not None:
        try:
            f = drive.CreateFile({"id": file_id})
            f.FetchMetadata(fields="id,title,trashed")
            return "Missing" if f.get("trashed") else "Available"
        except Exception as exc:
            if "404" in str(exc).lower():
                return "Missing"
            logger.warning("Drive API error for ID '%s': %s — falling back to HTTP.", file_id, exc)
    return _public_gdrive_check(file_id, is_folder=is_folder)


# ═══════════════════════════════════════════════════════════════════════════════
#  GOOGLE DRIVE DOWNLOADER
# ═══════════════════════════════════════════════════════════════════════════════

def download_from_gdrive(
    file_id:       str,
    dest_dir:      str,
    name:          str,
    drive:         Optional[object] = None,
    original_link: str = "",
) -> tuple[bool, str]:
    """
    Download a Google Drive file **or folder** to *dest_dir*.

    Folders (``/folders/`` URLs) are downloaded recursively via ``gdown``.
    Individual files fall back to the pydrive2 API or a direct HTTP download.
    """
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)
    is_folder = "/folders/" in original_link

    # ── Folder download via gdown ─────────────────────────────────────────────
    if is_folder:
        if not GDOWN_AVAILABLE:
            return False, "gdown not installed — run: pip install gdown"
        folder_url = f"https://drive.google.com/drive/folders/{file_id}"
        safe       = re.sub(r'[<>:"/\\|?*]+', "_", name).strip("_") or "download"
        out_dir    = str(dest / safe)
        try:
            gdown.download_folder(
                url=folder_url,
                output=out_dir,
                quiet=True,
                use_cookies=False,
            )
            return True, out_dir
        except Exception as exc:
            return False, f"gdown folder error: {exc}"

    # ── Single-file download ──────────────────────────────────────────────────
    if drive is not None:
        try:
            gf   = drive.CreateFile({"id": file_id})
            gf.FetchMetadata(fields="id,title,mimeType")
            safe = re.sub(r'[<>:"/\\|?*\s]+', "_", gf.get("title", name)).strip("_") or "download"
            out  = dest / safe
            gf.GetContentFile(str(out))
            return True, str(out)
        except Exception as exc:
            return False, f"pydrive2 error: {exc}"

    try:
        safe = re.sub(r'[<>:"/\\|?*\s]+', "_", name).strip("_") or "download"
        url  = f"https://drive.google.com/uc?id={file_id}&export=download&confirm=t"
        req  = urllib.request.Request(url)
        req.add_header("User-Agent", "Mozilla/5.0 (AuditChecker/1.0)")
        with urllib.request.urlopen(req, timeout=60) as resp:
            ct = resp.headers.get("Content-Type", "")
            if "text/html" in ct:
                return False, "Requires authentication (large/restricted file)"
            cd  = resp.headers.get("Content-Disposition", "")
            m   = re.search(r'filename[^;=\n]*=["\']?([^"\';\n]+)', cd)
            out = dest / (m.group(1).strip() if m else safe)
            with open(out, "wb") as fh:
                fh.write(resp.read())
        return True, str(out)
    except Exception as exc:
        return False, f"HTTP error: {exc}"


# ═══════════════════════════════════════════════════════════════════════════════
#  DISK SPACE UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def get_disk_usage(path: str) -> tuple[float, float, float]:
    """
    Return *(total_gb, used_gb, free_gb)* for the filesystem that contains
    *path*.  Walks up to the nearest existing ancestor if *path* itself does
    not yet exist (e.g. a download destination that hasn't been created yet).
    """
    try:
        p = Path(path)
        while not p.exists() and p.parent != p:
            p = p.parent
        usage = shutil.disk_usage(str(p))
        gb = 1024 ** 3
        return usage.total / gb, usage.used / gb, usage.free / gb
    except Exception as exc:
        logger.warning("Could not read disk usage for '%s': %s", path, exc)
        return 0.0, 0.0, float("inf")


def print_disk_stats(path: str, label: str = "") -> None:
    """Print a one-line disk usage bar for the filesystem at *path*."""
    try:
        total, used, free = get_disk_usage(path)
        if total == 0:
            print("      Disk: (unknown)")
            return
        pct    = used / total * 100
        filled = int(30 * used / total)
        bar    = "\u2588" * filled + "\u2591" * (30 - filled)
        prefix = f"  [{label}] " if label else "      "
        print(
            f"{prefix}Disk [{bar}] {pct:.1f}% used  |  "
            f"{free:.2f} GB free / {total:.2f} GB total"
        )
    except Exception:
        pass


# ═══════════════════════════════════════════════════════════════════════════════
#  REPORT GENERATION
# ═══════════════════════════════════════════════════════════════════════════════

def generate_report(results: list[CourseAuditResult], output_path: str) -> None:
    """
    Write a colour-coded .xlsx report.

    Column layout:
        Course | Semester | Term | Status |
        <AssetKey>_Local | <AssetKey>_Local_Path | <AssetKey>_Drive  (× 6 assets)

    Row colour:
        Green  — course folder found locally AND all available Drive links are live
        Yellow — partially found (some local or some Drive links OK)
        Red    — nothing found anywhere
    """
    asset_order = list(ASSET_COLUMNS.keys())

    records = []
    for r in results:
        row: dict = {
            "Course":   r.course_name,
            "Semester": r.semester,
            "Term":     r.term,
            "Status":   r.status,
        }
        for col_name in asset_order:
            k  = ASSET_COLUMNS[col_name]["key"]
            ar = r.asset_results.get(col_name, AssetResult())
            row[f"{k}_Local"]      = ar.found_locally
            row[f"{k}_Local_Path"] = ar.local_path
            row[f"{k}_Drive"]      = ar.drive_status
        records.append(row)

    df  = pd.DataFrame(records)
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(out), engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Audit Report")
        ws = writer.sheets["Audit Report"]

        # Auto-size columns
        for col_cells in ws.columns:
            width = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(width + 4, 60)

        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        fill_green  = PatternFill("solid", fgColor="C6EFCE")   # all present
        fill_yellow = PatternFill("solid", fgColor="FFEB9C")   # partial
        fill_red    = PatternFill("solid", fgColor="FFC7CE")   # nothing found

        # Column positions for scoring
        col_names      = list(df.columns)
        local_indices  = [i for i, c in enumerate(col_names) if c.endswith("_Local")]
        drive_indices  = [i for i, c in enumerate(col_names) if c.endswith("_Drive")]

        for row in ws.iter_rows(min_row=2):
            local_yes  = sum(1 for i in local_indices  if str(row[i].value or "").startswith("Yes"))
            drive_avail = sum(1 for i in drive_indices if str(row[i].value or "") == "Available")
            total_local = len(local_indices)
            total_drive = sum(1 for i in drive_indices if str(row[i].value or "") not in {"No Link", "Not Checked"})

            if local_yes == total_local and (total_drive == 0 or drive_avail == total_drive):
                fill = fill_green
            elif local_yes == 0 and drive_avail == 0:
                fill = fill_red
            else:
                fill = fill_yellow

            for cell in row:
                cell.fill = fill

    logger.info("Report written to: %s", output_path)
    print(f"\n  Report saved → {output_path}")


# ═══════════════════════════════════════════════════════════════════════════════
#  AUDIT ORCHESTRATION
# ═══════════════════════════════════════════════════════════════════════════════

def run_audit(
    excel_dir:            str,
    drives:               list[str],
    output:               str,
    course_col:           str,
    sem_col:              str,
    term_col:             str,
    status_col:           str,
    fuzzy_threshold:      int,
    cache_file:           str,
    cache_max_age_hours:  float,
    max_workers:          int,
    extensions_filter:    Optional[list[str]],
    gdrive_enabled:       bool,
    credentials_file:     str,
    settings_file:        str,
    download_missing:     bool = False,
    download_dest:        Optional[str] = None,
    download_all:         bool = False,
    min_free_gb:          float = 5.0,
) -> None:

    # ── Step 1: Read Excel ────────────────────────────────────────────────────
    print("\n[1/4] Reading Excel / CSV input files …")
    courses = read_excel_courses(excel_dir, course_col, sem_col, term_col, status_col)
    if not courses:
        logger.error("No course entries found — nothing to audit.")
        sys.exit(1)
    total_links = sum(len(c.asset_links) for c in courses)
    print(f"      {len(courses)} unique courses loaded.")
    print(f"      {total_links} Drive hyperlinks extracted across all asset columns.")

    # ── Step 2: Scan local drives ─────────────────────────────────────────────
    all_paths: list[str] = []
    if drives:
        print(f"\n[2/4] Scanning {len(drives)} drive(s) …")
        _, all_paths = build_file_index(
            drives,
            cache_file=cache_file,
            cache_max_age_hours=cache_max_age_hours,
            max_workers=max_workers,
            extensions_filter=extensions_filter,
        )
        print(f"      {len(all_paths):,} total paths indexed (files + folders).")
    else:
        print("\n[2/4] No drives specified — skipping local scan.")

    # ── Step 3: Authenticate Google Drive ────────────────────────────────────
    gdrive = None
    if gdrive_enabled:
        print("\n[3/4] Connecting to Google Drive …")
        gdrive = authenticate_gdrive(credentials_file, settings_file)
        print("      " + ("Authenticated via pydrive2." if gdrive else "Using public HTTP fallback."))
    else:
        print("\n[3/4] Google Drive link checking disabled (pass --gdrive to enable).")

    # ── Step 4: Audit every course ────────────────────────────────────────────
    print(f"\n[4/4] Auditing {len(courses)} course(s) across {len(ASSET_COLUMNS)} asset types …")
    results: list[CourseAuditResult] = []

    for course in tqdm(courses, desc="Auditing", unit="course"):
        result = CourseAuditResult(
            course_name=course.course_name,
            semester=course.semester,
            term=course.term,
            status=course.status,
        )

        # Find all local paths matching this course name — done ONCE per course
        # then reused cheaply for all six asset columns.
        candidates = match_course_paths(course.course_name, all_paths, fuzzy_threshold) if all_paths else []

        for col_name, col_info in ASSET_COLUMNS.items():
            ar = AssetResult()

            # ── Local check ───────────────────────────────────────────────────
            found, path = find_asset_in_candidates(candidates, col_info["hints"])
            if found:
                ar.found_locally = "Yes"
                ar.local_path    = path

            # ── Google Drive check ────────────────────────────────────────────
            link = course.asset_links.get(col_name)
            if link and gdrive_enabled:
                ar.drive_status = check_gdrive_link(link, gdrive)
            elif link:
                ar.drive_status = "Link Present (not checked)"
            else:
                ar.drive_status = "No Link"

            result.asset_results[col_name] = ar

        results.append(result)

    # ── Step 5: Download (per-course, all-or-nothing to avoid cross-disk splits) ─
    downloaded_ok = downloaded_fail = skipped_courses = 0
    do_download = download_missing or download_all

    if do_download:
        if not download_dest:
            print("\n[5/4] --download / --download_all requires --download_dest — skipping.")
        else:
            Path(download_dest).mkdir(parents=True, exist_ok=True)

            # ── Build a per-course map of assets to download ───────────────────
            # Key = course_name  →  list of (CourseAuditResult, CourseEntry, col_name)
            course_assets: dict[str, list[tuple]] = {}
            for r, c in zip(results, courses):
                for col_name, ar in r.asset_results.items():
                    link = c.asset_links.get(col_name)
                    if not link:
                        continue
                    if download_all and ar.drive_status not in (
                        "No Link", "Broken Link", "Missing"
                    ):
                        course_assets.setdefault(c.course_name, []).append((r, c, col_name))
                    elif download_missing and ar.found_locally == "No" and ar.drive_status == "Available":
                        course_assets.setdefault(c.course_name, []).append((r, c, col_name))

            total_assets = sum(len(v) for v in course_assets.values())

            if course_assets:
                mode = "all linked" if download_all else "missing"
                print(
                    f"\n[5/4] {total_assets} {mode} asset(s) across "
                    f"{len(course_assets)} course(s) → '{download_dest}'"
                )
                print_disk_stats(download_dest, "Before")
                print(
                    f"      Min free space : {min_free_gb:.1f} GB  "
                    f"(whole course is skipped if space is low — no cross-disk splits)"
                )

                for course_name, assets in tqdm(
                    course_assets.items(), desc="Courses", unit="course"
                ):
                    # ── Disk-space guard — check BEFORE touching any file ──────
                    _, _, free_gb = get_disk_usage(download_dest)
                    if free_gb < min_free_gb:
                        tqdm.write(
                            f"\n  LOW SPACE: {free_gb:.2f} GB free "
                            f"(< {min_free_gb:.1f} GB threshold) — "
                            f"skipping '{course_name}' to keep course files together."
                        )
                        skipped_courses += 1
                        continue

                    # ── Download every asset for this course ───────────────────
                    for result, course, col_name in assets:
                        link    = course.asset_links.get(col_name)
                        file_id = extract_drive_file_id(link) if link else None
                        if not file_id:
                            continue

                        # Folder structure: <dest>/<Course Name>/<Asset Type>/
                        safe_course = re.sub(r'[<>:"/\\|?*]+', "_", course.course_name).strip()
                        asset_key   = ASSET_COLUMNS[col_name]["key"]
                        course_dest = str(Path(download_dest) / safe_course / asset_key)

                        ok, msg = download_from_gdrive(
                            file_id, course_dest, asset_key, gdrive,
                            original_link=link or "",
                        )
                        ar = result.asset_results[col_name]
                        if ok:
                            ar.found_locally = "Yes (Downloaded)"
                            ar.local_path    = msg
                            downloaded_ok   += 1
                        else:
                            logger.warning(
                                "Download failed [%s / %s]: %s",
                                course.course_name, col_name, msg,
                            )
                            downloaded_fail += 1

                print_disk_stats(download_dest, "After ")
            else:
                mode = "all linked" if download_all else "missing"
                print(f"\n[5/4] No {mode} assets with Drive links — nothing to download.")
    else:
        print("\n[5/4] Download step skipped (use --download or --download_all to enable).")

    # ── Generate report ───────────────────────────────────────────────────────
    generate_report(results, output)

    # ── Summary ───────────────────────────────────────────────────────────────
    n             = len(results)
    fully_local   = sum(
        1 for r in results
        if all(ar.found_locally.startswith("Yes") for ar in r.asset_results.values())
    )
    partial_local = sum(
        1 for r in results
        if any(ar.found_locally.startswith("Yes") for ar in r.asset_results.values())
        and not all(ar.found_locally.startswith("Yes") for ar in r.asset_results.values())
    )
    none_local    = sum(
        1 for r in results
        if not any(ar.found_locally.startswith("Yes") for ar in r.asset_results.values())
    )
    has_links = sum(1 for c in courses if c.asset_links)

    # ── Unmatched courses detail ───────────────────────────────────────────────
    unmatched = [
        (r, c) for r, c in zip(results, courses)
        if not any(ar.found_locally.startswith("Yes") for ar in r.asset_results.values())
    ]
    if unmatched:
        print("\n  Courses with NO local match (check folder names on your drive):")
        print("  " + "-" * 74)
        for result, course in unmatched:
            link_count = len(course.asset_links)
            artifacts  = course.asset_links.get("Course Artifacts Link", "—")
            print(f"  • {result.course_name}")
            print(f"    Links in Excel : {link_count}/6")
            if artifacts and artifacts != "—":
                print(f"    Drive folder   : {artifacts}")
        print("  " + "-" * 74)

    dl_lines = ""
    if do_download:
        dl_lines = (
            f"|  Downloaded OK        : {downloaded_ok:>8,}    |\n"
            f"|  Download failures    : {downloaded_fail:>8,}    |\n"
            f"|  Courses skipped (low space): {skipped_courses:>3,}    |\n"
        )

    print(f"""
+-------------------------------------------+
|            AUDIT SUMMARY                  |
+-------------------------------------------+
|  Total courses          : {n:>8,}    |
|  Courses with links     : {has_links:>8,}    |
|  All assets found local : {fully_local:>8,}    |
|  Some assets found local: {partial_local:>8,}    |
|  No local assets found  : {none_local:>8,}    |
{dl_lines}+-------------------------------------------+
""")


# ═══════════════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════════════

def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python3 audit_checker.py",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description="""
╔══════════════════════════════════════════════════════════════════╗
║                  COURSERA AUDIT CHECKER                         ║
║  Cross-check Google Drive course folders against local drives   ║
╚══════════════════════════════════════════════════════════════════╝

Reads a Google Sheet (or local Excel file) that lists courses and
their Google Drive folder links, scans your connected hard drives
for matching content, checks whether each Drive folder is still
accessible, and produces a colour-coded Excel report.

Default behaviour (no flags needed if config.json is set up):
  • Downloads the Google Sheet from the URL in config.json
  • Scans the drives listed in config.json
  • Checks every Drive link (Available / Missing / Broken)
  • Saves a report to ./availability_report.xlsx
""",
        epilog="""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 FIRST-TIME SETUP  (do this once after cloning)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  Step 1 — Install dependencies:
       pip install -r requirements.txt

  Step 2 — Edit config.json and set your values:
       "gsheet_url"  →  paste your Google Sheets link
       "drives"      →  paths to your connected hard drives
                        e.g. ["/run/media/yourname/One Touch A"]
       "download_dest" → drive where missing files get saved

  Step 3 — Run the audit:
       ./audit_checker.py

  That's it. The sheet is fetched automatically, drives are scanned,
  every Drive link is checked, and a report is saved.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 COMMON WORKFLOWS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

  1. Standard audit  (URL + drives already in config.json):
       python3 audit_checker.py

  2. Audit then download everything missing onto the drive:
       python3 audit_checker.py --download

  3. Download missing files to a specific drive:
       python3 audit_checker.py --download --download_dest "/run/media/duffer/One Touch B"

  4. Force a fresh copy of the Google Sheet (ignore 1-hour cache):
       python3 audit_checker.py --no_cache

  5. Use a different Google Sheet URL just for this run:
       python3 audit_checker.py --gsheet_url "https://docs.google.com/spreadsheets/d/SPREADSHEET_ID/edit"

  6. Override which drives to scan:
       python3 audit_checker.py --drives "/run/media/duffer/One Touch A" "/run/media/duffer/One Touch B"

  7. Save the report to a custom location:
       python3 audit_checker.py --output "/home/duffer/Desktop/report.xlsx"

  8. Full run — fresh sheet + rescan drives + download missing files:
       python3 audit_checker.py --no_cache --download

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 REPORT COLOURS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  🟢 Green   — all assets found locally AND all Drive links are live
  🟡 Yellow  — some assets found locally or some Drive links are live
  🔴 Red     — nothing found locally and no Drive links are live

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 DRIVE STATUS VALUES IN THE REPORT
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Available          — folder exists and is publicly accessible
  Missing            — folder is private, deleted, or inaccessible
  Broken Link        — URL could not be parsed or request failed
  No Link            — no Google Drive link in the spreadsheet row

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
 CONFIG  (config.json — edit these instead of typing flags every run)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  gsheet_url          Google Sheets URL (primary input)
  drives              List of local hard-drive mount paths
  output              Report output path
  google_drive
    download_dest     Default download destination drive
    enabled           true = always check Drive links (recommended)
  scanning
    fuzzy_threshold   Name-match sensitivity 0-100 (default 75)
    gsheet_cache_hours  Re-download sheet after N hours (default 1)
    cache_max_age_hours Re-scan drives after N hours (default 24)
""",
    )

    # ── Input ──────────────────────────────────────────────────────────────────
    inp = parser.add_argument_group("INPUT  (where to read course data from)")
    inp.add_argument(
        "--gsheet_url", metavar="URL",
        help="Google Sheets URL to fetch and audit. "
             "The sheet must be shared as 'Anyone with the link can view'. "
             "Takes precedence over --excel_dir. "
             "(default: value in config.json)",
    )
    inp.add_argument(
        "--excel_dir", metavar="DIR",
        help="Folder containing .xlsx / .csv course files. "
             "Used only when --gsheet_url is not set. "
             "(default: ./excel)",
    )
    inp.add_argument(
        "--config", metavar="FILE",
        help="Path to a custom config.json. "
             "(default: ./config.json)",
    )

    # ── Drives ─────────────────────────────────────────────────────────────────
    drv = parser.add_argument_group("DRIVES  (which local drives to scan)")
    drv.add_argument(
        "--drives", nargs="+", metavar="PATH",
        help="One or more local drive/folder paths to scan for course files. "
             "Example: --drives \"/run/media/duffer/One Touch A\" \"/run/media/duffer/One Touch B\" "
             "(default: drives list in config.json)",
    )

    # ── Output ─────────────────────────────────────────────────────────────────
    out = parser.add_argument_group("OUTPUT  (where to save the report)")
    out.add_argument(
        "--output", metavar="FILE",
        help="Path for the Excel audit report. "
             "(default: ./availability_report.xlsx)",
    )

    # ── Download ───────────────────────────────────────────────────────────────
    dl = parser.add_argument_group(
        "DOWNLOAD  (fetch missing assets from Google Drive onto local drives)"
    )
    dl.add_argument(
        "--download", action="store_true",
        help="After auditing, download every asset that is Available on "
             "Google Drive but not found locally. "
             "Files are saved as: <download_dest>/<Course Name>/<Asset Type>/",
    )
    dl.add_argument(
        "--download_dest", metavar="DIR",
        help="Root folder on the local drive where missing assets are saved. "
             "Example: --download_dest \"/run/media/duffer/One Touch A\" "
             "(default: google_drive.download_dest in config.json)",
    )
    dl.add_argument(
        "--download_all", action="store_true",
        help="Download ALL linked assets from Google Drive, even if already "
             "found locally (full sync). Like --download but ignores local presence. "
             "Files are saved as: <download_dest>/<Course Name>/<Asset Type>/",
    )
    dl.add_argument(
        "--min_free_gb", type=float, metavar="GB",
        help="Minimum free disk space (GB) to keep on the download drive. "
             "If free space drops below this before a course starts, the entire "
             "course is skipped so it is never split across disks. "
             "(default: 5.0 GB)",
    )

    # ── Cache ──────────────────────────────────────────────────────────────────
    cch = parser.add_argument_group("CACHE  (control re-downloading and re-scanning)")
    cch.add_argument(
        "--no_cache", action="store_true",
        help="Ignore ALL cached data and start fresh: "
             "re-downloads the Google Sheet AND re-scans all drives. "
             "Use this after the spreadsheet has been updated.",
    )

    # ── Advanced ───────────────────────────────────────────────────────────────
    adv = parser.add_argument_group("ADVANCED")
    adv.add_argument(
        "--gdrive", action="store_true",
        help="Force Drive link checking on even if not enabled in config.json. "
             "(Drive checking is already ON by default — this flag is only "
             "needed if you set enabled=false in config.json and want to "
             "override it for one run.)",
    )
    adv.add_argument(
        "--fuzzy_threshold", type=int, metavar="0-100",
        help="Minimum fuzzy-match score (0-100) for a local folder name to "
             "be considered a match for a course name. "
             "Lower = more permissive, higher = stricter. "
             "(default: 75)",
    )
    adv.add_argument(
        "--workers", type=int, metavar="N",
        help="Number of parallel workers used when scanning drives. "
             "(default: number of CPU cores)",
    )
    adv.add_argument(
        "--log_level", metavar="LEVEL",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Logging verbosity written to audit_checker.log. "
             "Choices: DEBUG | INFO | WARNING | ERROR  (default: INFO)",
    )

    return parser


def main() -> None:
    parser = _build_arg_parser()
    args   = parser.parse_args()

    config = load_config(args.config)

    gsheet_url  = getattr(args, "gsheet_url", None) or config.get("gsheet_url", "")
    excel_dir   = args.excel_dir       or config["excel_dir"]
    drives      = args.drives          or config["drives"]
    output      = args.output          or config["output"]
    fuzzy_thr   = args.fuzzy_threshold or config["scanning"]["fuzzy_threshold"]
    max_workers = args.workers         or config["scanning"]["max_workers"]
    log_level   = args.log_level       or config["logging"]["level"]
    cache_file    = config["scanning"]["cache_file"]
    cache_age     = 0.0 if args.no_cache else float(config["scanning"]["cache_max_age_hours"])
    gsheet_cache  = float(config["scanning"].get("gsheet_cache_hours", 1))
    ext_filter    = config["scanning"].get("extensions_filter") or []
    gdrive_on     = args.gdrive or config["google_drive"].get("enabled", False)
    creds_file    = config["google_drive"]["credentials_file"]
    settings_f    = config["google_drive"]["settings_file"]
    min_free_gb   = (
        args.min_free_gb
        if getattr(args, "min_free_gb", None) is not None
        else float(config["google_drive"].get("min_free_gb", 5.0))
    )
    log_file      = config["logging"]["log_file"]
    course_col    = config["columns"].get("course_name", "Course")
    sem_col       = config["columns"].get("semester",    "Sem")
    term_col      = config["columns"].get("term",        "Term")
    status_col    = config["columns"].get("status",      "Status")

    global logger
    logger = setup_logging(log_file, log_level)

    # ── Download Google Sheet (primary input source) ───────────────────────────
    if gsheet_url:
        # Download into a dedicated cache folder so it does not mix with any
        # manually placed Excel files in excel_dir.
        gsheet_cache_dir = "./gsheet_cache"
        print("\n[0/4] Fetching Google Sheet …")
        downloaded = download_gsheet(
            gsheet_url,
            dest_dir=gsheet_cache_dir,
            cache_hours=gsheet_cache,
            force=args.no_cache,
        )
        if not downloaded:
            logger.error("Could not download Google Sheet — aborting.")
            sys.exit(1)
        # Override excel_dir so only the downloaded sheet is read.
        excel_dir = gsheet_cache_dir
    elif not any(Path(excel_dir).rglob("*.xlsx")) and not any(Path(excel_dir).rglob("*.csv")):
        print(f"\nERROR: No input source found.")
        print(f"  Set 'gsheet_url' in config.json  OR  place .xlsx/.csv files in '{excel_dir}'.")
        sys.exit(1)

    valid_drives: list[str] = []
    for d in drives:
        p = Path(d)
        if p.exists():
            valid_drives.append(str(p.resolve()))
        else:
            logger.warning("Drive path does not exist, skipping: %s", d)

    logger.info("=== Audit Checker starting ===")
    logger.info("Excel dir  : %s", excel_dir)
    logger.info("Drives     : %s", valid_drives)
    logger.info("Output     : %s", output)
    logger.info("Fuzzy thr  : %d", fuzzy_thr)
    logger.info("GDrive on  : %s", gdrive_on)

    run_audit(
        excel_dir=excel_dir,
        drives=valid_drives,
        output=output,
        course_col=course_col,
        sem_col=sem_col,
        term_col=term_col,
        status_col=status_col,
        fuzzy_threshold=fuzzy_thr,
        cache_file=cache_file,
        cache_max_age_hours=cache_age,
        max_workers=max_workers,
        extensions_filter=ext_filter or None,
        gdrive_enabled=gdrive_on,
        credentials_file=creds_file,
        settings_file=settings_f,
        download_missing=args.download,
        download_dest=args.download_dest or config["google_drive"].get("download_dest", ""),
        download_all=getattr(args, "download_all", False),
        min_free_gb=min_free_gb,
    )


# ═══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    multiprocessing.freeze_support()
    main()
