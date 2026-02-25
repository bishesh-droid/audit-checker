#!/usr/bin/env python3
"""
downloader.py
=============
Downloads course assets from Google Drive (links stored in the master Google
Sheet), organises them course-by-course on two external drives, and produces
a colour-coded Excel progress report.

Rules
-----
* Each course lives entirely on ONE disk (no splitting across drives).
* If a course folder already exists on a disk, that disk is used.
* New courses go to the disk with more free space at the time of assignment.
* Downloads are resumable: already-downloaded folders are skipped.

Usage
-----
    python3 downloader.py                  # full run
    python3 downloader.py --report-only    # regenerate Excel without downloading
    python3 downloader.py --course "Discrete Mathematics"   # one course only
    python3 downloader.py --refresh-sheet  # re-download the Google Sheet first
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import shutil
import sys
import time
import urllib.request
from pathlib import Path
from typing import Optional

# ── Third-party ───────────────────────────────────────────────────────────────
import subprocess
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────
# rclone remote name (configured via `rclone config`, name it "gdrive")
RCLONE_REMOTE = "gdrive"

SHEET_ID   = "1Kb7AcEmVZDLg5lgV6pHJQ8lE40sJMIL0"
SHEET_URL  = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SHEET_CACHE = Path(__file__).parent / "gsheet_cache" / "master_sheet.xlsx"

DISKS = [
    Path("/run/media/duffer/One Touch"),
    Path("/run/media/duffer/One Touch A"),
]

# Root subfolder on each disk where downloaded courses will be placed
COURSE_ROOT = "Downloaded Courses"

# Where the progress/report Excel will be written (next to this script)
PROGRESS_FILE = Path(__file__).parent / "download_progress.xlsx"

# State file that records which disk each course is assigned to
STATE_FILE = Path(__file__).parent / "disk_assignment.json"

LOG_FILE = Path(__file__).parent / "downloader.log"

# Google Sheet tab to read
SHEET_TAB = "RCA"

# Column indices (0-based) in the RCA sheet
COL_COURSE   = 0
COL_SEM      = 1
COL_TERM     = 2
COL_FACULTY  = 3
COL_LOCATION = 4
COL_VENDOR   = 5
COL_CO       = 6   # Course Outline
COL_PPTS     = 7   # PPTs
COL_WA       = 8   # Written Assets
COL_FV       = 9   # Final Videos
COL_RV       = 10  # Raw Videos
COL_STATUS   = 13  # Status
COL_CA       = 12  # Course Artifacts Link

ASSET_COLS = {
    "Course Outline":  COL_CO,
    "PPTs":            COL_PPTS,
    "Written Assets":  COL_WA,
    "Final Videos":    COL_FV,
    "Raw Videos":      COL_RV,
    "Course Artifacts": COL_CA,
}

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_name(name: str) -> str:
    """Make a course name safe for use as a folder name."""
    name = name.strip()
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name[:120]  # Cap length


def extract_folder_id(url: str) -> Optional[str]:
    """Extract Google Drive folder ID from various URL formats."""
    if not url:
        return None
    patterns = [
        r'/folders/([a-zA-Z0-9_-]{25,})',
        r'id=([a-zA-Z0-9_-]{25,})',
        r'/d/([a-zA-Z0-9_-]{25,})',
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None


def free_bytes(path: Path) -> int:
    """Return free bytes on the disk containing path."""
    try:
        stat = shutil.disk_usage(path)
        return stat.free
    except Exception:
        return 0


def human_size(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024:
            return f"{n:.1f} {unit}"
        n /= 1024
    return f"{n:.1f} PB"


# ═══════════════════════════════════════════════════════════════════════════════
#  GOOGLE SHEET LOADING
# ═══════════════════════════════════════════════════════════════════════════════

def download_sheet(force: bool = False) -> Path:
    """Download the Google Sheet xlsx to the cache. Returns the cache path."""
    SHEET_CACHE.parent.mkdir(parents=True, exist_ok=True)
    if not force and SHEET_CACHE.exists():
        age_hours = (time.time() - SHEET_CACHE.stat().st_mtime) / 3600
        if age_hours < 1:
            log.info("Using cached sheet (%.0f min old)", age_hours * 60)
            return SHEET_CACHE
    log.info("Downloading master sheet …")
    try:
        req = urllib.request.Request(SHEET_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as r:
            data = r.read()
        SHEET_CACHE.write_bytes(data)
        log.info("Sheet downloaded (%d bytes)", len(data))
    except Exception as exc:
        if SHEET_CACHE.exists():
            log.warning("Sheet download failed (%s); using cached copy.", exc)
        else:
            raise RuntimeError(f"Cannot download sheet and no cache: {exc}") from exc
    return SHEET_CACHE


def parse_courses(xlsx_path: Path) -> list[dict]:
    """Parse the RCA sheet and return a list of course dicts with Drive links."""
    from openpyxl import load_workbook as _lw
    wb = _lw(xlsx_path, data_only=True)

    if SHEET_TAB not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_TAB}' not found in {xlsx_path}")

    ws = wb[SHEET_TAB]
    courses = []

    for row in ws.iter_rows(min_row=2):
        course_name = row[COL_COURSE].value
        if not course_name or not str(course_name).strip():
            continue

        status = str(row[COL_STATUS].value or "").strip()
        links = {}
        for asset_name, col_idx in ASSET_COLS.items():
            cell = row[col_idx]
            href = cell.hyperlink.target if cell.hyperlink else None
            links[asset_name] = href  # may be None

        courses.append({
            "course":  str(course_name).strip(),
            "status":  status,
            "links":   links,
        })

    log.info("Parsed %d courses from sheet", len(courses))
    return courses


# ═══════════════════════════════════════════════════════════════════════════════
#  DISK ASSIGNMENT
# ═══════════════════════════════════════════════════════════════════════════════

def load_state() -> dict:
    """Load the disk-assignment state file (course → disk index)."""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            pass
    return {}


def save_state(state: dict) -> None:
    STATE_FILE.write_text(json.dumps(state, indent=2, ensure_ascii=False))


def find_existing_disk(course_name: str) -> Optional[int]:
    """Return the index (0 or 1) of the disk that already has this course folder."""
    safe = sanitize_name(course_name)
    for idx, disk in enumerate(DISKS):
        folder = disk / COURSE_ROOT / safe
        if folder.exists() and any(folder.iterdir()):
            return idx
    return None


def assign_disk(course_name: str, state: dict) -> int:
    """
    Return the disk index to use for this course.
    Priority: existing folder > state file > disk with most free space.
    """
    # 1. Does a course folder already exist on disk?
    existing = find_existing_disk(course_name)
    if existing is not None:
        return existing

    # 2. Already in state (assigned but not yet downloaded)
    if course_name in state:
        return state[course_name]

    # 3. Pick disk with most free space
    free = [free_bytes(d) for d in DISKS]
    chosen = free.index(max(free))
    log.info(
        "Assigning '%s' → Disk %d (%s) [free: %s vs %s]",
        course_name, chosen,
        DISKS[chosen].name,
        human_size(free[chosen]), human_size(free[1 - chosen]),
    )
    return chosen


# ═══════════════════════════════════════════════════════════════════════════════
#  DOWNLOADING
# ═══════════════════════════════════════════════════════════════════════════════

def folder_is_populated(path: Path) -> bool:
    """Return True if the folder exists and contains at least one file."""
    if not path.exists():
        return False
    for item in path.rglob("*"):
        if item.is_file():
            return True
    return False


def download_drive_folder(folder_url: str, dest: Path, retries: int = 3) -> bool:
    """
    Download a Google Drive folder to dest using rclone.
    Returns True on success, False on failure.

    Requires rclone to be configured with a remote named RCLONE_REMOTE
    (default: "gdrive"). Run `rclone config` once to set it up.
    """
    folder_id = extract_folder_id(folder_url)
    if not folder_id:
        log.warning("  Cannot parse folder ID from: %s", folder_url)
        return False

    dest.mkdir(parents=True, exist_ok=True)

    # rclone copy with --drive-root-folder-id downloads the contents of the
    # specified folder ID into dest, preserving subfolder structure.
    cmd = [
        "rclone", "copy",
        f"{RCLONE_REMOTE}:",          # remote (root overridden by flag below)
        str(dest),
        f"--drive-root-folder-id={folder_id}",
        "--progress",
        "--transfers=4",              # parallel file transfers
        "--checkers=8",
        "--retries=3",
        "--low-level-retries=10",
        "--stats=10s",
    ]

    for attempt in range(1, retries + 1):
        try:
            log.info("  rclone: folder=%s → %s (attempt %d)", folder_id, dest, attempt)
            result = subprocess.run(cmd, capture_output=False, text=True)
            if result.returncode == 0:
                log.info("  rclone completed successfully")
                return True
            log.warning("  rclone exit code %d on attempt %d", result.returncode, attempt)
        except Exception as exc:
            log.warning("  Attempt %d error: %s", attempt, exc)

        if attempt < retries:
            wait = 15 * attempt
            log.info("  Retrying in %ds …", wait)
            time.sleep(wait)

    log.error("  All %d attempts failed for folder %s", retries, folder_id)
    return False


def download_course(course: dict, disk_idx: int) -> dict:
    """
    Download all asset folders for a course to the assigned disk.

    Each asset type gets its own subfolder:
        <course_dir>/Course Outline/
        <course_dir>/PPTs/
        <course_dir>/Written Assets/
        <course_dir>/Final Videos/
        <course_dir>/Raw Videos/
        <course_dir>/Course Artifacts/   ← production/umbrella folder

    Returns a results dict: {asset_name: 'ok'|'skipped'|'no_link'|'failed'}
    """
    course_name = course["course"]
    safe = sanitize_name(course_name)
    course_dir = DISKS[disk_idx] / COURSE_ROOT / safe
    course_dir.mkdir(parents=True, exist_ok=True)

    results = {}

    # Download every asset link into its own named subfolder.
    # Duplicate Drive IDs (same folder linked in multiple columns) are
    # detected and the second occurrence is marked 'skipped' to avoid
    # re-downloading identical content.
    seen_ids: set[str] = set()

    for asset_name, asset_url in course["links"].items():
        if not asset_url:
            results[asset_name] = "no_link"
            continue

        folder_id = extract_folder_id(asset_url)
        asset_dir = course_dir / asset_name

        # Already downloaded in a previous run?
        if folder_is_populated(asset_dir):
            log.info("  '%s / %s' already present – skipping", course_name, asset_name)
            results[asset_name] = "skipped"
            if folder_id:
                seen_ids.add(folder_id)
            continue

        # Duplicate link in this course (same Drive folder ID used twice)?
        if folder_id and folder_id in seen_ids:
            log.info("  '%s / %s' is a duplicate link – skipping", course_name, asset_name)
            results[asset_name] = "skipped"
            continue

        log.info("  Downloading '%s / %s' …", course_name, asset_name)
        ok = download_drive_folder(asset_url, asset_dir)
        results[asset_name] = "ok" if ok else "failed"
        if folder_id and ok:
            seen_ids.add(folder_id)

    return results


# ═══════════════════════════════════════════════════════════════════════════════
#  PROGRESS REPORT
# ═══════════════════════════════════════════════════════════════════════════════

# Colour palette
GREEN  = PatternFill("solid", fgColor="70AD47")   # downloaded / ok
YELLOW = PatternFill("solid", fgColor="FFD966")   # skipped (already present)
RED    = PatternFill("solid", fgColor="FF7575")   # failed
ORANGE = PatternFill("solid", fgColor="F4B942")   # no link / in production
GREY   = PatternFill("solid", fgColor="BFBFBF")   # not started
BLUE   = PatternFill("solid", fgColor="9DC3E6")   # header
WHITE  = PatternFill("solid", fgColor="FFFFFF")

BOLD = Font(bold=True)
BOLD_WHITE = Font(bold=True, color="FFFFFF")


def status_fill(val: str, course_status: str) -> PatternFill:
    if course_status == "In Production":
        return ORANGE
    v = str(val).lower()
    if v in ("ok",):
        return GREEN
    if v in ("skipped",):
        return YELLOW
    if v in ("failed",):
        return RED
    if v in ("no_link",):
        return ORANGE
    if v in ("not started",):
        return GREY
    return WHITE


def build_report(
    courses: list[dict],
    state: dict,
    download_results: dict[str, dict],
) -> None:
    """Write the Excel progress report to PROGRESS_FILE."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Download Progress"

    asset_names = list(ASSET_COLS.keys())  # ordered
    headers = (
        ["#", "Course", "Status", "Assigned Disk", "Disk Path"]
        + asset_names
        + ["Overall"]
    )

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = BLUE
        cell.font = BOLD_WHITE
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 2, 18)

    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["E"].width = 55
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_num, course in enumerate(courses, 2):
        name   = course["course"]
        status = course["status"]
        disk_i = state.get(name)
        disk_label = DISKS[disk_i].name if disk_i is not None else "—"
        disk_path  = (
            str(DISKS[disk_i] / COURSE_ROOT / sanitize_name(name))
            if disk_i is not None
            else "—"
        )

        res = download_results.get(name, {})

        row_data = [row_num - 1, name, status, disk_label, disk_path]

        # Per-asset columns
        asset_vals = []
        for an in asset_names:
            v = res.get(an, "not started" if status == "Completed" else "n/a")
            asset_vals.append(v)
        row_data.extend(asset_vals)

        # Overall status
        if status == "In Production":
            overall = "In Production"
        elif not res:
            overall = "Not Started"
        elif all(v in ("ok", "skipped") for v in res.values()):
            overall = "Complete"
        elif any(v in ("ok", "skipped") for v in res.values()):
            overall = "Partial"
        elif all(v == "failed" for v in res.values()):
            overall = "Failed"
        else:
            overall = "Pending"
        row_data.append(overall)

        # Write cells
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            # Colour asset columns
            header = headers[col_idx - 1]
            if header in asset_names:
                cell.fill = status_fill(val, status)
            elif header == "Overall":
                ov = str(val).lower()
                if ov == "complete":
                    cell.fill = GREEN
                elif ov == "partial":
                    cell.fill = YELLOW
                elif ov in ("failed",):
                    cell.fill = RED
                elif ov in ("in production", "n/a"):
                    cell.fill = ORANGE
                elif ov == "not started":
                    cell.fill = GREY

    # Legend sheet
    leg = wb.create_sheet("Legend")
    legend_items = [
        ("Green",       "70AD47", "Downloaded / OK"),
        ("Yellow",      "FFD966", "Already present / Skipped"),
        ("Red",         "FF7575", "Download failed"),
        ("Orange",      "F4B942", "No link or In Production"),
        ("Grey",        "BFBFBF", "Not yet started"),
    ]
    for i, (label, colour, meaning) in enumerate(legend_items, 1):
        leg.cell(i, 1, label).fill = PatternFill("solid", fgColor=colour)
        leg.cell(i, 2, meaning)

    wb.save(PROGRESS_FILE)
    log.info("Progress report saved → %s", PROGRESS_FILE)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--report-only", action="store_true",
                   help="Only (re)generate the Excel report without downloading")
    p.add_argument("--refresh-sheet", action="store_true",
                   help="Force re-download the Google Sheet even if cache is fresh")
    p.add_argument("--course", metavar="NAME", action="append", dest="courses",
                   help="Download only courses whose name contains this string (case-insensitive). "
                        "Repeat the flag to download multiple: --course 'Discrete' --course 'Writing'")
    p.add_argument("--dry-run", action="store_true",
                   help="Show what would be downloaded without actually downloading")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    # ── Load sheet ────────────────────────────────────────────────────────────
    xlsx = download_sheet(force=args.refresh_sheet)
    courses = parse_courses(xlsx)

    # ── Filter by --course (supports multiple flags) ──────────────────────────
    if args.courses:
        needles = [n.lower() for n in args.courses]
        courses = [
            c for c in courses
            if any(needle in c["course"].lower() for needle in needles)
        ]
        if not courses:
            log.error("No courses matched: %s", args.courses)
            sys.exit(1)
        log.info("Filtered to %d course(s) matching %s", len(courses), args.courses)

    # ── Check disks ───────────────────────────────────────────────────────────
    for disk in DISKS:
        if not disk.exists():
            log.error("Disk not found: %s  (is it plugged in?)", disk)
            sys.exit(1)
        (disk / COURSE_ROOT).mkdir(parents=True, exist_ok=True)

    # ── Load state (disk assignments) ─────────────────────────────────────────
    state = load_state()

    # ── Assign disks for all courses (balanced across both disks) ────────────
    # Track how many new courses are assigned to each disk so we can alternate
    new_assignments = [0, 0]  # count of new courses assigned to disk 0 and 1
    for course in courses:
        name = course["course"]
        if name not in state:
            existing = find_existing_disk(name)
            if existing is not None:
                state[name] = existing
            else:
                # Balance: prefer disk with most free space, but if both have
                # similar space, alternate assignments to spread load evenly
                free = [free_bytes(d) for d in DISKS]
                # If one disk has >20% more free space, use it; otherwise alternate
                if free[0] > free[1] * 1.20:
                    chosen = 0
                elif free[1] > free[0] * 1.20:
                    chosen = 1
                else:
                    # Alternate to balance load
                    chosen = 0 if new_assignments[0] <= new_assignments[1] else 1
                new_assignments[chosen] += 1
                log.info(
                    "Assigning '%s' → Disk %d (%s) [free: %s | %s]",
                    name, chosen, DISKS[chosen].name,
                    human_size(free[0]), human_size(free[1]),
                )
                state[name] = chosen
    save_state(state)

    # ── Load existing download results ────────────────────────────────────────
    results_file = Path(__file__).parent / "download_results.json"
    if results_file.exists():
        try:
            download_results: dict[str, dict] = json.loads(results_file.read_text())
        except Exception:
            download_results = {}
    else:
        download_results = {}

    if args.report_only:
        log.info("--report-only: skipping downloads, regenerating report …")
        build_report(courses, state, download_results)
        return

    # ── Download loop ─────────────────────────────────────────────────────────
    completable = [c for c in courses if c["status"] == "Completed" and any(c["links"].values())]
    in_prod     = [c for c in courses if c["status"] == "In Production"]
    no_links    = [c for c in courses if c["status"] == "Completed" and not any(c["links"].values())]

    log.info("─" * 60)
    log.info("Courses to download : %d", len(completable))
    log.info("In Production (skip): %d", len(in_prod))
    log.info("Completed/no links  : %d", len(no_links))
    log.info("─" * 60)

    for i, course in enumerate(completable, 1):
        name   = course["course"]
        disk_i = state[name]

        log.info("[%d/%d] %s  →  %s", i, len(completable), name, DISKS[disk_i].name)

        if args.dry_run:
            log.info("  [DRY RUN] Would download to: %s", DISKS[disk_i] / COURSE_ROOT / sanitize_name(name))
            continue

        try:
            res = download_course(course, disk_i)
        except KeyboardInterrupt:
            log.warning("Interrupted by user – saving progress …")
            break
        except Exception as exc:
            log.error("Unexpected error downloading '%s': %s", name, exc)
            res = {k: "failed" for k in ASSET_COLS}

        download_results[name] = res
        results_file.write_text(json.dumps(download_results, indent=2, ensure_ascii=False))

        ok_count   = sum(1 for v in res.values() if v in ("ok", "skipped"))
        fail_count = sum(1 for v in res.values() if v == "failed")
        log.info("  Result: %d ok/skipped, %d failed", ok_count, fail_count)
        log.info("  Disk free after: %s", human_size(free_bytes(DISKS[disk_i])))
        log.info("")

    # ── Build report ──────────────────────────────────────────────────────────
    build_report(courses, state, download_results)
    log.info("All done.")


if __name__ == "__main__":
    main()
